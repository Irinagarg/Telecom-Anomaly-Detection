# app.py
from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import numpy as np
import io
import os
import tempfile
import threading
import uuid
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pyxlsb

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024  # 500MB

# ── KPI Config ──
KPI_CONFIG = [
    {
        "col"      : "ERAB Drop Rate_Radio_Nokia_VeMr",
        "label"    : "Drop Rate",
        "direction": "high",
        "threshold": 0.5,
    },
    {
        "col"      : "E2E Call Setup Success Rate_Nokia_VeMr",
        "label"    : "E2E CSSR",
        "direction": "low",
        "threshold": 99.0,
    },
    {
        "col"      : "VoLTE Call Setup Success Rate_Nokia_VeMr",
        "label"    : "VoLTE CSSR",
        "direction": "low",
        "threshold": 99.0,
    },
    {
        "col"      : "Average CQI_Nokia_VeMr",
        "label"    : "CQI",
        "direction": "low",
        "threshold": 7.0,
    },
    {
        "col"      : "Nokia_LTE_DL_User_Throughput_Mbps",
        "label"    : "Avg Thpr",
        "direction": "low",
        "threshold": 3.0,
    },
]

CELL_COL = "CO_DN"

# ── Styles ──
TTL_FILL  = PatternFill("solid", fgColor="0D2137")
TTL_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=12)
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL  = PatternFill("solid", fgColor="D6E4F0")
WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")
RED_FILL  = PatternFill("solid", fgColor="FFC7CE")
GRN_FILL  = PatternFill("solid", fgColor="E2EFDA")
BODY_FONT = Font(name="Arial", size=9)
RED_FONT  = Font(name="Arial", size=9, color="C00000", bold=True)
GRN_FONT  = Font(name="Arial", size=9, color="375623")
GRY_FONT  = Font(name="Arial", size=9, color="AAAAAA")
BLK_FONT  = Font(name="Arial", size=9, color="000000")
BLD_RED   = Font(name="Arial", size=9, color="C00000", bold=True)
BLD_BLK   = Font(name="Arial", size=9, bold=True)
THIN      = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── In-memory job store ──
jobs = {}


def short_cell_name(co_dn):
    parts = str(co_dn).split("/")
    if len(parts) >= 2:
        return f"{parts[-2]}_{parts[-1]}"
    return str(co_dn)


def write_title_header(ws, title, headers):
    ncols = len(headers)
    ws.append([title] + [""] * (ncols - 1))
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"].fill      = TTL_FILL
    ws["A1"].font      = TTL_FONT
    ws["A1"].alignment = CENTER
    ws["A1"].border    = BORDER
    ws.append(headers)
    for cell in ws[2]:
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28


def apply_base_style(ws, start_row, end_row, ncols):
    for i in range(start_row, end_row + 1):
        fill = ALT_FILL if i % 2 == 0 else None
        for j in range(1, ncols + 1):
            c = ws.cell(row=i, column=j)
            c.font      = BODY_FONT
            c.border    = BORDER
            c.alignment = LEFT if j == 1 else CENTER
            if fill:
                c.fill = fill


def process_file(tmp_path):
    """Read sheets, auto-slim, track all values + failures."""

    # ── Detect file type ──
    try:
        if tmp_path.endswith(".xlsb"):
            with pyxlsb.open_workbook(tmp_path) as wb:
                sheet_names = wb.sheets
            engine = "pyxlsb"
        else:
            xl          = pd.ExcelFile(tmp_path, engine="openpyxl")
            sheet_names = xl.sheet_names
            engine      = "openpyxl"
    except Exception as e:
        raise Exception(f"Could not read file: {e}")

    # ── AUTO SLIM ──
    KEEP_COLS = ["CO_DN"] + [cfg["col"] for cfg in KPI_CONFIG]
    print(f"Auto-slimming {len(sheet_names)} sheets...")

    slimmed_sheets = {}
    for sheet in sheet_names:
        try:
            if engine == "pyxlsb":
                df = pd.read_excel(
                    tmp_path, sheet_name=sheet, engine="pyxlsb"
                )
            else:
                df = pd.read_excel(
                    tmp_path, sheet_name=sheet, engine="openpyxl"
                )
        except Exception as e:
            print(f"Skipping {sheet}: {e}")
            continue

        if df.empty or len(df) < 50:
            print(f"Skipping {sheet} — only {len(df)} rows")
            continue

        keep     = [c for c in KEEP_COLS if c in df.columns]
        slim_df  = df[keep].copy()
        orig_mb  = df.memory_usage(deep=True).sum() / 1024 / 1024
        slim_mb  = slim_df.memory_usage(deep=True).sum() / 1024 / 1024
        print(f"  {sheet}: {len(df)} rows, "
              f"{len(df.columns)} cols → {len(keep)} cols "
              f"({orig_mb:.1f}MB → {slim_mb:.1f}MB)")
        slimmed_sheets[sheet] = slim_df
        del df

    if not slimmed_sheets:
        return None, None, None, None, None, None

    print(f"Processing {len(slimmed_sheets)} sheets...")

    failure_records = []
    all_values      = []
    valid_sheets    = []
    all_cells_seen  = set()

    for sheet, df_day in slimmed_sheets.items():
        df_day = df_day.reset_index(drop=True)

        if CELL_COL not in df_day.columns:
            new_header     = df_day.iloc[0]
            df_day         = df_day[1:].reset_index(drop=True)
            df_day.columns = new_header
            if CELL_COL not in df_day.columns:
                print(f"Skipping {sheet} — no {CELL_COL}")
                continue

        valid_sheets.append(sheet)
        df_day["Short_Cell"] = df_day[CELL_COL].apply(short_cell_name)
        all_cells_seen.update(df_day["Short_Cell"].unique())

        for cfg in KPI_CONFIG:
            col = cfg["col"]
            if col not in df_day.columns:
                continue

            numeric_col = pd.to_numeric(df_day[col], errors="coerce")

            if cfg["direction"] == "high":
                breach_mask = numeric_col > cfg["threshold"]
            else:
                breach_mask = numeric_col < cfg["threshold"]

            breach_mask = breach_mask.fillna(False)

            # ── Track ALL values (passing + failing) ──
            all_vals_df = pd.DataFrame({
                "Short_Cell": df_day["Short_Cell"].values,
                "KPI"       : cfg["label"],
                "Day"       : sheet,
                "Value"     : numeric_col.round(3).values,
                "Failing"   : breach_mask.values,
            })
            all_values.append(all_vals_df)

            # ── Track only failures for summary ──
            if breach_mask.sum() > 0:
                failed_df = pd.DataFrame({
                    "Short_Cell": df_day.loc[breach_mask, "Short_Cell"].values,
                    "KPI"       : cfg["label"],
                    "Day"       : sheet,
                    "Value"     : numeric_col[breach_mask].round(3).values,
                    "Threshold" : cfg["threshold"],
                    "Direction" : cfg["direction"],
                })
                failure_records.append(failed_df)

        del df_day

    print(f"Valid sheets    : {valid_sheets}")
    print(f"Total cells seen: {len(all_cells_seen)}")
    sheet_names   = valid_sheets

    if not failure_records:
        return None, None, None, None, None, None

    failure_df   = pd.concat(failure_records, ignore_index=True)
    all_values_df = pd.concat(all_values,     ignore_index=True) \
                   if all_values else pd.DataFrame()

    # ── KPI Pivot ──
    kpi_failure_counts = (
        failure_df.groupby(["Short_Cell", "KPI"])
        .size()
        .reset_index(name="Days_Failed")
    )
    kpi_pivot = kpi_failure_counts.pivot_table(
        index      = "Short_Cell",
        columns    = "KPI",
        values     = "Days_Failed",
        fill_value = 0
    ).reset_index()
    kpi_pivot.columns.name = None

    kpi_labels       = [cfg["label"] for cfg in KPI_CONFIG]
    kpi_cols_present = [c for c in kpi_labels if c in kpi_pivot.columns]

    kpi_pivot["Total_KPI_Failures"] = kpi_pivot[kpi_cols_present].sum(axis=1)
    kpi_pivot["KPIs_Ever_Failed"]   = (kpi_pivot[kpi_cols_present] > 0).sum(axis=1)
    kpi_pivot["Max_Days_Any_KPI"]   = kpi_pivot[kpi_cols_present].max(axis=1)
    kpi_pivot = kpi_pivot.sort_values(
        "Total_KPI_Failures", ascending=False
    ).reset_index(drop=True)

    # ── KPI Summary ──
    max_days         = len(sheet_names)
    kpi_summary_rows = []

    for cfg in KPI_CONFIG:
        label    = cfg["label"]
        kpi_data = failure_df[failure_df["KPI"] == label]

        if kpi_data.empty:
            kpi_summary_rows.append({
                "KPI"                   : label,
                "Threshold"             : cfg["threshold"],
                "Cells_Ever_Failed"     : 0,
                "Total_Failure_Events"  : 0,
                "Avg_Days_Failed"       : 0,
                "Max_Days_By_One_Cell"  : 0,
                "Cells_Failed_All_Days" : 0,
                "Worst_Cell"            : "—",
                "Worst_Cell_Days"       : 0,
            })
            continue

        per_cell = (
            kpi_data.groupby("Short_Cell")
            .agg(
                days_failed = ("Day",   "count"),
                avg_value   = ("Value", "mean"),
            )
            .reset_index()
        )

        if cfg["direction"] == "high":
            per_cell = per_cell.sort_values(
                ["days_failed", "avg_value"],
                ascending=[False, False]
            )
        else:
            per_cell = per_cell.sort_values(
                ["days_failed", "avg_value"],
                ascending=[False, True]
            )

        worst_row = per_cell.iloc[0]
        kpi_summary_rows.append({
            "KPI"                   : label,
            "Threshold"             : cfg["threshold"],
            "Cells_Ever_Failed"     : per_cell["days_failed"].count(),
            "Total_Failure_Events"  : len(kpi_data),
            "Avg_Days_Failed"       : int(round(per_cell["days_failed"].mean(), 0)),
            "Max_Days_By_One_Cell"  : per_cell["days_failed"].max(),
            "Cells_Failed_All_Days" : len(per_cell[per_cell["days_failed"] == max_days]),
            "Worst_Cell"            : worst_row["Short_Cell"],
            "Worst_Cell_Days"       : int(worst_row["days_failed"]),
        })

    kpi_summary_df = pd.DataFrame(kpi_summary_rows)
    return (failure_df, kpi_pivot, kpi_summary_df,
            sheet_names, all_cells_seen, all_values_df)


def generate_excel(failure_df, kpi_pivot, kpi_summary_df,
                   sheet_names, all_cells_seen, all_values_df):
    """Generate Excel and return as bytes."""

    kpi_labels       = [cfg["label"] for cfg in KPI_CONFIG]
    kpi_cols_present = [c for c in kpi_labels if c in kpi_pivot.columns]
    max_days         = len(sheet_names)

    # Pre-build fill cache
    FILL_CACHE = {}
    for d in range(0, max_days + 1):
        if d == 0:
            FILL_CACHE[d] = WHT_FILL
        else:
            intensity     = int((d / max_days) * 200)
            green         = max(0, 255 - intensity)
            blue          = max(0, 255 - intensity)
            FILL_CACHE[d] = PatternFill(
                "solid", fgColor=f"FF{green:02X}{blue:02X}"
            )

    def days_fill(days):
        return FILL_CACHE.get(int(days), WHT_FILL)

    def days_font(days):
        if days == 0:
            return GRY_FONT
        if days >= max_days * 0.8:
            return RED_FONT
        return BLK_FONT

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: KPI Summary ──
    s1 = wb.create_sheet("KPI Summary")
    write_title_header(s1, "KPI Failure Summary — Vi PJB 4G", [
        "KPI", "Threshold", "Cells Ever Failed",
        "Total Failure Events", "Avg Days Failed",
        "Max Days (1 Cell)", "Cells Failed All Days",
        "Worst Cell", "Worst Cell Days"
    ])
    for _, row in kpi_summary_df.iterrows():
        s1.append([
            row["KPI"],
            row["Threshold"],
            row["Cells_Ever_Failed"],
            row["Total_Failure_Events"],
            row["Avg_Days_Failed"],
            row["Max_Days_By_One_Cell"],
            row["Cells_Failed_All_Days"],
            row["Worst_Cell"],
            row["Worst_Cell_Days"],
        ])
    apply_base_style(s1, 3, 3 + len(kpi_summary_df) - 1, 9)
    for col, w in zip("ABCDEFGHI", [16, 12, 18, 20, 16, 18, 20, 28, 16]):
        s1.column_dimensions[col].width = w
    s1.freeze_panes = "A3"

    # ── Sheet 2: Failure Matrix ──
    s2 = wb.create_sheet("Failure Matrix")
    matrix_headers = (
        ["Cell ID (eNodeB_Cell)"] + kpi_cols_present +
        ["Total Failure Days", "KPIs Ever Failed", "Max Days Any KPI"]
    )
    write_title_header(
        s2,
        "KPI Failure Matrix — Days Failed per Cell per KPI",
        matrix_headers
    )

    pivot_vals = kpi_pivot[kpi_cols_present].values.astype(int)
    short_vals = kpi_pivot["Short_Cell"].values
    total_vals = kpi_pivot["Total_KPI_Failures"].astype(int).values
    ever_vals  = kpi_pivot["KPIs_Ever_Failed"].astype(int).values
    max_vals   = kpi_pivot["Max_Days_Any_KPI"].astype(int).values

    for idx in range(len(kpi_pivot)):
        s2.append(
            [short_vals[idx]] +
            list(pivot_vals[idx]) +
            [total_vals[idx], ever_vals[idx], max_vals[idx]]
        )

    apply_base_style(s2, 3, 3 + len(kpi_pivot) - 1, len(matrix_headers))

    n_kpis = len(kpi_cols_present)
    for idx in range(len(kpi_pivot)):
        row_i = idx + 3
        for k in range(n_kpis):
            val  = pivot_vals[idx][k]
            cell = s2.cell(row=row_i, column=k + 2)
            cell.fill = days_fill(val)
            cell.font = days_font(val)
        total_col = n_kpis + 2
        s2.cell(row=row_i, column=total_col).font = (
            BLD_RED if total_vals[idx] >= max_days * 2 else BLD_BLK
        )

    s2.column_dimensions["A"].width = 24
    for idx in range(2, len(matrix_headers) + 1):
        s2.column_dimensions[get_column_letter(idx)].width = 16
    s2.freeze_panes = "B3"

    # ── Sheets 3–7: One per KPI — ALL cells + ALL values ──
    all_cells_df = pd.DataFrame({
        "Short_Cell": sorted(list(all_cells_seen))
    })

    for cfg in KPI_CONFIG:
        label    = cfg["label"]
        kpi_data = failure_df[failure_df["KPI"] == label]

        # All values for this KPI
        kpi_all  = all_values_df[
            all_values_df["KPI"] == label
        ] if not all_values_df.empty else pd.DataFrame()

        # Per cell failure stats
        if not kpi_data.empty:
            per_cell_fails = (
                kpi_data.groupby("Short_Cell")
                .agg(
                    Days_Failed = ("Day",   "count"),
                    Avg_Value   = ("Value", "mean"),
                    Min_Value   = ("Value", "min"),
                    Max_Value   = ("Value", "max"),
                )
                .reset_index()
            )
        else:
            per_cell_fails = pd.DataFrame(columns=[
                "Short_Cell", "Days_Failed",
                "Avg_Value", "Min_Value", "Max_Value"
            ])

        # Per cell overall stats from ALL values
        if not kpi_all.empty:
            per_cell_all = (
                kpi_all.groupby("Short_Cell")
                .agg(
                    Avg_Value_All = ("Value", "mean"),
                    Min_Value_All = ("Value", "min"),
                    Max_Value_All = ("Value", "max"),
                )
                .reset_index()
            )
        else:
            per_cell_all = pd.DataFrame(columns=[
                "Short_Cell",
                "Avg_Value_All", "Min_Value_All", "Max_Value_All"
            ])

        # Merge ALL cells
        per_cell = all_cells_df.merge(
            per_cell_fails, on="Short_Cell", how="left"
        ).merge(
            per_cell_all, on="Short_Cell", how="left"
        )

        per_cell["Days_Failed"]   = per_cell["Days_Failed"].fillna(0).astype(int)
        per_cell["Persistence_%"] = (
            per_cell["Days_Failed"] / max_days * 100
        ).round(1)

        # Use overall avg/min/max (not just failing values)
        per_cell["Avg_Value"] = per_cell["Avg_Value_All"].fillna(0).round(3)
        per_cell["Min_Value"] = per_cell["Min_Value_All"].fillna(0).round(3)
        per_cell["Max_Value"] = per_cell["Max_Value_All"].fillna(0).round(3)

        # Sort — failing cells first
        if cfg["direction"] == "high":
            per_cell = per_cell.sort_values(
                ["Days_Failed", "Avg_Value"],
                ascending=[False, False]
            ).reset_index(drop=True)
        else:
            per_cell = per_cell.sort_values(
                ["Days_Failed", "Avg_Value"],
                ascending=[False, True]
            ).reset_index(drop=True)

        ws = wb.create_sheet(title=label[:31])
        kpi_headers = (
            ["Cell ID (eNodeB_Cell)", "Days Failed", "Persistence %",
             f"Avg {label}", f"Min {label}", f"Max {label}"]
            + [s[:12] for s in sheet_names]
        )

        n_failing = len(per_cell_fails)
        n_total   = len(per_cell)
        write_title_header(
            ws,
            f"{label} — Threshold "
            f"{'>' if cfg['direction'] == 'high' else '<'}{cfg['threshold']}  "
            f"({n_failing} failing cells  |  {n_total} total cells)",
            kpi_headers
        )

        # Day pivot — ALL values
        if not kpi_all.empty:
            day_pivot = kpi_all.pivot_table(
                index   = "Short_Cell",
                columns = "Day",
                values  = "Value",
                aggfunc = "first"
            ).reindex(columns=sheet_names)

            # Failing flag pivot
            failing_pivot = kpi_all.pivot_table(
                index   = "Short_Cell",
                columns = "Day",
                values  = "Failing",
                aggfunc = "first"
            ).reindex(columns=sheet_names).fillna(False)
        else:
            day_pivot     = pd.DataFrame(index=[], columns=sheet_names)
            failing_pivot = pd.DataFrame(index=[], columns=sheet_names)

        days_failed_list = []
        is_failing_list  = []

        for idx in range(len(per_cell)):
            p_row   = per_cell.iloc[idx]
            cell_id = p_row["Short_Cell"]
            df_val  = int(p_row["Days_Failed"])
            days_failed_list.append(df_val)
            is_failing_list.append(df_val > 0)

            ws_row = [
                cell_id,
                df_val,
                f"{p_row['Persistence_%']}%",
                round(float(p_row["Avg_Value"]), 3),
                round(float(p_row["Min_Value"]), 3),
                round(float(p_row["Max_Value"]), 3),
            ]

            # Day columns — show ALL values
            if cell_id in day_pivot.index:
                for v in day_pivot.loc[cell_id]:
                    ws_row.append(
                        round(float(v), 3) if pd.notna(v) else ""
                    )
            else:
                ws_row += [""] * len(sheet_names)

            ws.append(ws_row)

        n_rows = len(per_cell)
        n_cols = len(kpi_headers)
        apply_base_style(ws, 3, 3 + n_rows - 1, n_cols)

        for idx in range(n_rows):
            row_i      = idx + 3
            df_val     = days_failed_list[idx]
            is_failing = is_failing_list[idx]
            cell_id    = per_cell.iloc[idx]["Short_Cell"]

            # Days Failed column colour
            if is_failing:
                ws.cell(row=row_i, column=2).fill = days_fill(df_val)
                ws.cell(row=row_i, column=2).font = days_font(df_val)
            else:
                ws.cell(row=row_i, column=2).fill = GRN_FILL
                ws.cell(row=row_i, column=2).font = GRN_FONT

            # Day value columns — red if failing, normal if passing
            if cell_id in day_pivot.index:
                day_vals = day_pivot.loc[cell_id].values
                fail_vals = (
                    failing_pivot.loc[cell_id].values
                    if cell_id in failing_pivot.index
                    else [False] * len(sheet_names)
                )
                for k, (v, is_fail) in enumerate(
                    zip(day_vals, fail_vals)
                ):
                    if pd.notna(v):
                        c = ws.cell(row=row_i, column=7 + k)
                        if is_fail:
                            c.fill = RED_FILL
                            c.font = RED_FONT

        ws.column_dimensions["A"].width = 24
        for idx in range(2, 7):
            ws.column_dimensions[get_column_letter(idx)].width = 14
        for idx in range(7, 7 + len(sheet_names)):
            ws.column_dimensions[get_column_letter(idx)].width = 12
        ws.freeze_panes = "G3"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ── Routes ──

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filename = file.filename.lower()
    if filename.endswith(".xlsb"):
        suffix = ".xlsb"
    elif filename.endswith(".xlsx"):
        suffix = ".xlsx"
    elif filename.endswith(".xls"):
        suffix = ".xls"
    else:
        return jsonify({"error": "Unsupported file. Use .xlsb or .xlsx"}), 400

    job_id       = str(uuid.uuid4())
    jobs[job_id] = {"status": "processing", "result": None, "error": None}

    tmp      = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp_path = tmp.name

    try:
        chunk_size = 1024 * 1024
        with open(tmp_path, "wb") as f:
            while True:
                chunk = file.stream.read(chunk_size)
                if not chunk:
                    break
                f.write(chunk)
    except Exception as e:
        return jsonify({"error": f"Upload failed: {e}"}), 500

    def run():
        try:
            result = process_file(tmp_path)
            if result[0] is None:
                jobs[job_id]["status"] = "error"
                jobs[job_id]["error"]  = (
                    "No KPI failures found — check column names"
                )
                return
            (failure_df, kpi_pivot, kpi_summary_df,
             sheet_names, all_cells_seen, all_values_df) = result

            excel_bytes = generate_excel(
                failure_df, kpi_pivot, kpi_summary_df,
                sheet_names, all_cells_seen, all_values_df
            )
            jobs[job_id]["status"] = "done"
            jobs[job_id]["result"] = excel_bytes
        except Exception as e:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"]  = str(e)
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    threading.Thread(target=run).start()
    return jsonify({"job_id": job_id})


@app.route("/status/<job_id>")
def status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({"status": job["status"], "error": job["error"]})


@app.route("/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "Not ready"}), 404

    excel_bytes = job["result"]
    del jobs[job_id]

    return send_file(
        io.BytesIO(excel_bytes),
        mimetype      = "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
        as_attachment = True,
        download_name = "KPI_Failure_Tracker.xlsx"
    )


@app.route("/health")
def health():
    return "OK", 200


@app.route("/debug", methods=["POST"])
def debug():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400

    file     = request.files["file"]
    suffix   = ".xlsx" if file.filename.endswith(".xlsx") else ".xlsb"
    tmp      = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp_path = tmp.name
    file.save(tmp_path)
    tmp.close()

    try:
        if suffix == ".xlsx":
            all_sheets = pd.read_excel(
                tmp_path, sheet_name=None, engine="openpyxl"
            )
        else:
            all_sheets = {}
            with pyxlsb.open_workbook(tmp_path) as wb:
                for sheet in wb.sheets:
                    all_sheets[sheet] = pd.read_excel(
                        tmp_path, sheet_name=sheet, engine="pyxlsb"
                    )

        result = {}
        for sheet, df in all_sheets.items():
            result[sheet] = {
                "rows"    : len(df),
                "has_codn": CELL_COL in df.columns,
                "cols_10" : list(df.columns[:10]),
            }
        return jsonify(result)

    except Exception as e:
        return jsonify({
            "error": str(e),
            "trace": traceback.format_exc()
        })
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)